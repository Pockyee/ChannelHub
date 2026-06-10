"""ChannelHub — 报表解析 ETL：MinIO 已备份 .eml → 附件识别 → raw.sell_through_<来源>。

两条解析路径(本文件按「共用 / Expert / Hutt / 编排」四段组织,互不纠缠):
  · Expert 渠道周报  .xlsx —— openpyxl 读表头,与表头签名注册表比对
      命中 → 逐行带血缘 INSERT ON CONFLICT 进 raw.sell_through_expert(幂等)
  · Hutt 网店订单    .csv  —— Shopify 订单导出(orders_export_*.csv)
      表头签名命中 → 同样带血缘幂等入 raw.sell_through_hutt_shop_de
  · 任一附件签名不命中 → 经 SMTP 给 ALERT_EMAIL_TO 发告警(同一文件只告警一次)

所有取值原样 TEXT 落 raw,不做规范化(解析归一在 core/mart 层)。
解析入库后**链式刷新 mart 物化层**(mart.refresh_all():单事务 TRUNCATE+INSERT
重建 dim_company/dim_store/dim_product/fact_*),保证一解析完 BI 即新鲜。
Hutt 的 BI 口径层(mart.v_hutt_shop_orders,见 009 迁移)是普通视图,无需刷新。
"""

from __future__ import annotations

import csv
import io
import os
import re
import smtplib
import email as email_lib
from datetime import date, datetime
from email.header import decode_header, make_header
from email.message import EmailMessage

import psycopg
from minio import Minio
from openpyxl import load_workbook
from prefect import flow, get_run_logger, task
from prefect.runtime import flow_run


# ===========================================================================
# 一、共用基础设施:env / MinIO / Postgres / 告警 / 通用工具
#     (Expert 与 Hutt 两条解析路径都只依赖这一段,彼此零依赖)
# ===========================================================================
def _norm(h) -> str:
    """表头归一:去掉非字母数字、转小写 —— 'Financial Status' → 'financialstatus'。"""
    return re.sub(r"[^0-9a-z]", "", str(h).strip().lower()) if h is not None else ""


def _env(name: str, default: str = "") -> str:
    return os.environ.get(name, default).strip()


def _decode(v) -> str:
    try:
        return str(make_header(decode_header(v))) if v else ""
    except Exception:
        return str(v)


def _minio() -> Minio:
    return Minio(
        _env("MINIO_ENDPOINT", "minio:9000"),
        access_key=_env("MINIO_ACCESS_KEY"),
        secret_key=_env("MINIO_SECRET_KEY"),
        secure=_env("MINIO_SECURE", "false").lower() == "true",
    )


def _pg():
    return psycopg.connect(
        host=_env("POSTGRES_HOST", "postgres"),
        port=int(_env("POSTGRES_PORT", "5432") or "5432"),
        dbname=_env("POSTGRES_DB", "channelhub"),
        user=_env("POSTGRES_USER"),
        password=_env("POSTGRES_PASSWORD"),
    )


def _send_alert(subject: str, body: str) -> None:
    host = _env("SMTP_HOST", "smtp.ionos.de")
    port = int(_env("SMTP_PORT", "465") or "465")
    user = _env("SMTP_USER")
    pw = _env("SMTP_PASSWORD")
    to = _env("ALERT_EMAIL_TO")
    if not (user and pw and to):
        raise RuntimeError("SMTP_USER/SMTP_PASSWORD/ALERT_EMAIL_TO 未配置，无法发告警")
    msg = EmailMessage()
    msg["From"] = user
    msg["To"] = to
    msg["Subject"] = subject
    msg.set_content(body)
    with smtplib.SMTP_SSL(host, port, timeout=30) as s:
        s.login(user, pw)
        s.send_message(msg)


def _maybe_alert(object_key, fn, reason, header_seen, subject, sender, stats, logger):
    """未识别附件：未告警过才发邮件并记录（去重）。"""
    with _pg() as conn:
        with conn.cursor() as cur:
            cur.execute(
                "SELECT 1 FROM raw.ingest_alert WHERE source_object_key=%s AND source_file_name=%s",
                (object_key, fn),
            )
            if cur.fetchone():
                stats["alert_suppressed"] += 1
                logger.info("未识别附件 %s 已告警过，跳过", fn)
                return
    body = (
        f"未能识别报表附件，未入库。\n\n"
        f"附件名: {fn}\n原因: {reason}\n"
        f"邮件主题: {subject}\n发件人: {sender}\n"
        f"MinIO 对象: {object_key}\n\n"
        f"读到的表头:\n{header_seen}\n\n"
        f"请确认该文件格式，或在对应解析段增加表头签名。"
    )
    _send_alert(f"[ChannelHub] 未识别的报表附件: {fn}", body)
    with _pg() as conn:
        with conn.cursor() as cur:
            cur.execute(
                "INSERT INTO raw.ingest_alert "
                "(source_object_key, source_file_name, reason, header_seen, email_subject, email_from) "
                "VALUES (%s,%s,%s,%s,%s,%s) "
                "ON CONFLICT (source_object_key, source_file_name) DO NOTHING",
                (object_key, fn, reason, header_seen, subject, sender),
            )
        conn.commit()
    stats["alerted"] += 1
    logger.warning("已发送未识别附件告警: %s", fn)


def _insert_rows(table, dbcols, object_key, fn, sheet, rows):
    """带血缘幂等批量入库(两条解析路径共用)。

    rows: 可迭代 (业务值…, msg_id, object_key, fn, sheet, rownum, run_id) 元组,
    业务值顺序与 dbcols 一致。冲突键 (source_object_key, source_sheet,
    source_row_number) DO NOTHING → 重放安全。返回 (inserted, skipped)。
    """
    insert_sql = (
        f"INSERT INTO {table} "
        f"({', '.join(dbcols)}, source_email_message_id, source_object_key, "
        f"source_file_name, source_sheet, source_row_number, ingestion_run_id) "
        f"VALUES ({', '.join(['%s'] * len(dbcols))}, %s, %s, %s, %s, %s, %s) "
        f"ON CONFLICT (source_object_key, source_sheet, source_row_number) DO NOTHING"
    )
    count_sql = (
        f"SELECT count(*) FROM {table} "
        f"WHERE source_object_key=%s AND source_file_name=%s AND source_sheet=%s"
    )
    cnt_args = (object_key, fn, sheet)

    attempted = 0
    batch = []
    with _pg() as conn:
        with conn.cursor() as cur:
            cur.execute(count_sql, cnt_args)
            before = cur.fetchone()[0]
            for r in rows:
                attempted += 1
                batch.append(r)
                if len(batch) >= 500:
                    cur.executemany(insert_sql, batch)
                    batch.clear()
            if batch:
                cur.executemany(insert_sql, batch)
            conn.commit()
            cur.execute(count_sql, cnt_args)
            after = cur.fetchone()[0]
    inserted = after - before
    skipped = attempted - inserted          # 已存在被 ON CONFLICT 跳过的（幂等可见）
    return inserted, skipped


# ===========================================================================
# 二、Expert 渠道周报解析(.xlsx,表头签名识别)
#     新增 xlsx 类供应商:在 XLSX_REGISTRY 加一条签名 + 对应 raw 表即可
# ===========================================================================
EXPERT_COLUMNS = [
    ("periodflag", "period_flag"),
    ("transactiondate", "transaction_date"),
    ("providername", "provider_name"),
    ("company", "company"),
    ("storeid", "store_id"),
    ("store", "store_name"),
    ("street", "street"),
    ("postalcode", "postal_code"),
    ("city", "city"),
    ("customerskucode", "customer_sku_code"),
    ("customerskuname", "customer_sku_name"),
    ("gtinbarcode", "gtin_barcode"),
    ("supplierskucode", "supplier_sku_code"),
    ("soldqtyoutlets", "sold_qty_outlets"),
    ("stockonhandqtyoutlets", "stock_on_hand_qty_outlets"),
]

XLSX_REGISTRY = {
    "expert": {
        "table": "raw.sell_through_expert",
        "columns": EXPERT_COLUMNS,
        "required": {n for n, _ in EXPERT_COLUMNS},
    },
    # 未来：'msd': {...}, 'telekom': {...}
}


def _cell_to_text(v):
    """Excel 单元格 → 忠实字符串；空 → None。不做业务转换。"""
    if v is None:
        return None
    if isinstance(v, bool):
        return "TRUE" if v else "FALSE"
    if isinstance(v, datetime):
        return v.strftime("%d.%m.%Y")          # 保持文件可见的德式日期形态
    if isinstance(v, date):
        return v.strftime("%d.%m.%Y")
    if isinstance(v, float):
        return str(int(v)) if v.is_integer() else repr(v)  # 不要科学计数/.0
    if isinstance(v, int):
        return str(v)
    s = str(v).strip()
    return s or None


def _find_header(ws):
    """在前 8 行内找表头行，返回 (1基行号, normalized列表, 原始列表) 或 (None,None,None)。"""
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=8, values_only=True), start=1):
        norms = [_norm(c) for c in row]
        if sum(1 for n in norms if n) >= 5:  # 至少 5 个非空表头才算候选
            return i, norms, list(row)
    return None, None, None


def _identify_xlsx(header_norms: list[str]):
    """返回 (supplier_key, spec) 或 (None, None)。"""
    present = {h for h in header_norms if h}
    for key, spec in XLSX_REGISTRY.items():
        if spec["required"] <= present:
            return key, spec
    return None, None


def _load_sheet(ws, header_row_idx, header_norms, spec, object_key, fn, msg_id, run_id):
    col_idx = {}
    for norm, dbcol in spec["columns"]:
        if norm in header_norms:
            col_idx[dbcol] = header_norms.index(norm)
    dbcols = list(col_idx.keys())

    def gen():
        rownum = header_row_idx
        for row in ws.iter_rows(min_row=header_row_idx + 1, values_only=True):
            rownum += 1
            vals = [_cell_to_text(row[col_idx[c]]) if col_idx[c] < len(row) else None
                    for c in dbcols]
            if all(v is None for v in vals):
                continue  # 整行空，跳过尾部空行
            yield tuple(vals) + (msg_id, object_key, fn, ws.title, rownum, run_id)

    return _insert_rows(spec["table"], dbcols, object_key, fn, ws.title, gen())


def _parse_expert_xlsx(payload, fn, object_key, msg_id, run_id,
                       subject, sender, stats, logger):
    """单个 .xlsx 附件:打开 → 各 sheet 找表头 → 签名识别 → 入库;不命中则告警。"""
    try:
        wb = load_workbook(io.BytesIO(payload), read_only=True, data_only=True)
    except Exception as e:
        logger.warning("无法打开 xlsx %s (%s): %s", fn, object_key, e)
        _maybe_alert(object_key, fn, "xlsx_open_error", str(e), subject, sender, stats, logger)
        return

    matched_any = False
    header_seen_repr = ""
    for ws in wb.worksheets:
        header_row_idx, header_norms, header_raw = _find_header(ws)
        if header_row_idx is None:
            continue
        supplier, spec = _identify_xlsx(header_norms)
        if not supplier:
            header_seen_repr = " | ".join(str(h) for h in header_raw if h is not None)
            continue
        matched_any = True
        ins, skp = _load_sheet(ws, header_row_idx, header_norms, spec,
                               object_key, fn, msg_id, run_id)
        stats["inserted"] += ins
        stats["skipped"] += skp
        logger.info("解析 %s / sheet=%s 供应商=%s → 入库 %d, 跳过(已存在) %d",
                    fn, ws.title, supplier, ins, skp)
    wb.close()

    if not matched_any:
        _maybe_alert(object_key, fn, "unrecognized_header", header_seen_repr,
                     subject, sender, stats, logger)


# ===========================================================================
# 三、Hutt Online Shop 订单解析(.csv,Shopify 订单导出 orders_export_*.csv)
#     与 Expert 完全平行:签名识别 → 带血缘幂等入 raw.sell_through_hutt_shop_de
# ===========================================================================
HUTT_TABLE = "raw.sell_through_hutt_shop_de"

# raw 表业务列(79 列,与 009 迁移建表一致;顺序即 Shopify 导出列序)。
# CSV 表头 → 列名规则:norm(表头) == 列名去下划线,仅 3 个例外(见下)。
_HUTT_DB_COLS = [
    "order_name", "email", "financial_status", "paid_at", "fulfillment_status",
    "fulfilled_at", "accepts_marketing", "currency", "subtotal", "shipping",
    "taxes", "total", "discount_code", "discount_amount", "shipping_method",
    "order_created_at", "lineitem_quantity", "lineitem_name", "lineitem_price",
    "lineitem_compare_at_price", "lineitem_sku", "lineitem_requires_shipping",
    "lineitem_taxable", "lineitem_fulfillment_status", "billing_name",
    "billing_street", "billing_address1", "billing_address2", "billing_company",
    "billing_city", "billing_zip", "billing_province", "billing_country",
    "billing_phone", "shipping_name", "shipping_street", "shipping_address1",
    "shipping_address2", "shipping_company", "shipping_city", "shipping_zip",
    "shipping_province", "shipping_country", "shipping_phone", "notes",
    "note_attributes", "cancelled_at", "payment_method", "payment_reference",
    "refunded_amount", "vendor", "outstanding_balance", "employee", "location",
    "device_id", "order_id", "tags", "risk_level", "source", "lineitem_discount",
    "tax_1_name", "tax_1_value", "tax_2_name", "tax_2_value", "tax_3_name",
    "tax_3_value", "tax_4_name", "tax_4_value", "tax_5_name", "tax_5_value",
    "phone", "receipt_number", "duties", "billing_province_name",
    "shipping_province_name", "payment_id", "payment_terms_name",
    "next_payment_due_at", "payment_references",
]

# Shopify 表头与列名不同形的例外:Name→order_name, Created at→order_created_at, Id→order_id
_HUTT_HEADER_ALIASES = {"order_name": "name", "order_created_at": "createdat", "order_id": "id"}

HUTT_COLUMNS = [(_HUTT_HEADER_ALIASES.get(c, c.replace("_", "")), c) for c in _HUTT_DB_COLS]

# 签名只取核心列子集:Shopify 不同版本导出会增减边缘列,核心列在才认定是订单导出
HUTT_REQUIRED = {
    "name", "email", "financialstatus", "currency", "total", "createdat",
    "lineitemquantity", "lineitemname", "lineitemprice", "lineitemsku",
}


def _parse_hutt_csv(payload, fn, object_key, msg_id, run_id,
                    subject, sender, stats, logger):
    """单个 .csv 附件:表头签名命中 Shopify 订单导出才入库,否则告警。

    血缘约定(与历史入库一致,保证重放幂等):source_sheet = 文件名(CSV 无
    sheet 概念),表头是第 1 行 → source_row_number 从 2 起。
    取值原样 TEXT、空串落 NULL(如 shipping_zip 的前导撇号 '3042 原样保留,
    清洗在 mart.v_hutt_shop_orders)。
    """
    text = payload.decode("utf-8-sig", errors="replace")  # Shopify 导出 UTF-8(或带 BOM)
    reader = csv.reader(io.StringIO(text))
    try:
        header = next(reader)
    except StopIteration:
        _maybe_alert(object_key, fn, "empty_csv", "", subject, sender, stats, logger)
        return
    header_norms = [_norm(h) for h in header]

    if not HUTT_REQUIRED <= {h for h in header_norms if h}:
        _maybe_alert(object_key, fn, "unrecognized_csv_header",
                     " | ".join(header), subject, sender, stats, logger)
        return

    col_idx = {dbcol: header_norms.index(norm)
               for norm, dbcol in HUTT_COLUMNS if norm in header_norms}
    dbcols = list(col_idx.keys())

    def gen():
        for rownum, row in enumerate(reader, start=2):
            vals = [(row[col_idx[c]].strip() or None) if col_idx[c] < len(row) else None
                    for c in dbcols]
            if all(v is None for v in vals):
                continue  # 整行空，跳过尾部空行
            yield tuple(vals) + (msg_id, object_key, fn, fn, rownum, run_id)

    ins, skp = _insert_rows(HUTT_TABLE, dbcols, object_key, fn, fn, gen())
    stats["inserted"] += ins
    stats["skipped"] += skp
    logger.info("解析 %s 来源=hutt_shop_de → 入库 %d, 跳过(已存在) %d", fn, ins, skp)


# ===========================================================================
# 四、邮件遍历与 flow 编排
#     按附件扩展名分发:.xlsx → Expert 段;.csv → Hutt 段;其余(签名图等)跳过
# ===========================================================================
@task(retries=2, retry_delay_seconds=20)
def list_eml_keys() -> list[str]:
    c = _minio()
    bucket = _env("MINIO_BUCKET", "email-archive")
    return [
        o.object_name
        for o in c.list_objects(bucket, prefix="email/", recursive=True)
        if o.object_name.endswith(".eml")
    ]


@task(retries=1, retry_delay_seconds=20)
def process_eml(object_key: str) -> dict:
    logger = get_run_logger()
    bucket = _env("MINIO_BUCKET", "email-archive")
    run_id = str(getattr(flow_run, "id", "") or "")
    stats = {"xlsx": 0, "csv": 0, "inserted": 0, "skipped": 0,
             "alerted": 0, "alert_suppressed": 0}

    raw = _minio().get_object(bucket, object_key).read()
    msg = email_lib.message_from_bytes(raw)
    msg_id = (msg.get("Message-ID") or "").strip()
    subject = _decode(msg.get("Subject"))
    sender = _decode(msg.get("From"))

    for part in msg.walk():
        fn = _decode(part.get_filename() or "")
        low = fn.lower()
        if not (low.endswith(".xlsx") or low.endswith(".csv")):
            continue  # 跳过签名内嵌图(image001/002)等非报表附件
        payload = part.get_payload(decode=True) or b""
        if not payload:
            continue
        if low.endswith(".xlsx"):
            stats["xlsx"] += 1
            _parse_expert_xlsx(payload, fn, object_key, msg_id, run_id,
                               subject, sender, stats, logger)
        else:
            stats["csv"] += 1
            _parse_hutt_csv(payload, fn, object_key, msg_id, run_id,
                            subject, sender, stats, logger)

    return stats


@task(retries=1, retry_delay_seconds=15)
def refresh_mart() -> dict:
    """解析入库后链式刷新 mart 物化层（单事务 TRUNCATE+INSERT 重建,返回各表行数）。"""
    logger = get_run_logger()
    with _pg() as conn:
        with conn.cursor() as cur:
            cur.execute(
                "SELECT dim_company, dim_store, dim_product, fact, fact_current "
                "FROM mart.refresh_all()"
            )
            r = cur.fetchone()
        conn.commit()
    res = {
        "dim_company": r[0], "dim_store": r[1], "dim_product": r[2],
        "fact": r[3], "fact_current": r[4],
    }
    logger.info("mart 刷新完成: %s", res)
    return res


@flow(name="parse-sell-through")
def parse_sell_through() -> dict:
    logger = get_run_logger()
    keys = list_eml_keys()
    logger.info("待扫描 .eml: %d", len(keys))

    total = {"xlsx": 0, "csv": 0, "inserted": 0, "skipped": 0,
             "alerted": 0, "alert_suppressed": 0}
    failed = 0
    for k in keys:
        try:
            r = process_eml(k)
            for kk in total:
                total[kk] += r[kk]
        except Exception as e:
            failed += 1
            logger.warning("处理 %s 失败: %s", k, e)

    logger.info("解析完成: %s, failed=%d", total, failed)

    # 链式刷新 mart：即便部分 .eml 失败,已入库 raw 也应反映到 BI；
    # 刷新失败 → flow 失败(陈旧 mart 是真问题),解析失败的告警随后再抛。
    total["mart"] = refresh_mart()

    if failed:
        raise RuntimeError(f"{failed} 个 .eml 处理失败，详见日志；其余已入库并刷新: {total}")
    return total


if __name__ == "__main__":
    parse_sell_through()
